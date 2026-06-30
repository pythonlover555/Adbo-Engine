"""Append matched profiles to data/matches.xlsx, deduped on profile_id.

The extension can re-scan the same queue across many runs, so writes must
be idempotent: we load the existing profile_ids, skip anything already
present, and append only new rows. openpyxl reads + writes the whole
workbook (fine for the thousands-of-rows scale this will ever hit).
"""
from __future__ import annotations

import threading
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from .config import DATA_DIR, MATCHES_XLSX
from .logging_setup import get_logger

log = get_logger("excel")

COLUMNS = [
    "profile_id",
    "name",
    "country",
    "decision",
    "funding_kind",
    "funding_usd",
    "revenue_signal",
    "reason",
    "matched_at",
    "profile_url",
]

# Serialise writes — FastAPI handlers can overlap and openpyxl is not
# concurrency-safe on the same file.
_lock = threading.Lock()


def _load_or_create() -> tuple[Any, Any]:
    if MATCHES_XLSX.is_file():
        wb = load_workbook(MATCHES_XLSX)
        return wb, wb.active
    wb = Workbook()
    ws = wb.active
    ws.title = "matches"
    ws.append(COLUMNS)
    return wb, ws


def _seen_ids(ws: Any) -> set[str]:
    ids: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]:
            ids.add(str(row[0]))
    return ids


def append_match(record: dict[str, Any]) -> bool:
    """Append one profile. Returns True if written, False if it was a
    duplicate (profile_id already present)."""
    profile_id = str(record.get("profile_id") or record.get("profile_url") or record.get("name") or "").strip()
    if not profile_id:
        log.warning("append_match: no usable id; skipping write")
        return False

    with _lock:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        wb, ws = _load_or_create()
        if profile_id in _seen_ids(ws):
            log.info("dedup: %s already in matches.xlsx", profile_id)
            return False

        ws.append([
            profile_id,
            record.get("name", ""),
            record.get("country", ""),
            record.get("decision", ""),
            record.get("funding_kind", ""),
            record.get("funding_usd", ""),
            record.get("revenue_signal", ""),
            record.get("reason", ""),
            record.get("matched_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            record.get("profile_url", ""),
        ])
        wb.save(MATCHES_XLSX)
        log.info("saved match %s (%s)", record.get("name", "?"), profile_id)
        return True


def count_matches() -> int:
    with _lock:
        if not MATCHES_XLSX.is_file():
            return 0
        wb = load_workbook(MATCHES_XLSX, read_only=True)
        ws = wb.active
        n = max(ws.max_row - 1, 0)  # minus header
        wb.close()
        return n
